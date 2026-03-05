import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from utils.metricas import calcular_metricas, calcular_qualidade_arquivo

def aplicar_estilo_cabecalho(sheet, columns_count):
    """Aplica estilo profissional ao cabeçalho da planilha"""
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    alignment = Alignment(horizontal='center', vertical='center')
    
    for col in range(1, columns_count + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment

def ajustar_largura_colunas(sheet):
    """Ajusta a largura das colunas com base no conteúdo"""
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = min(adjusted_width, 50)

def formatar_valores_excel(sheet, columns_monetarias, columns_percentuais):
    """Aplica formatos numéricos do Excel para valores monetários e percentuais"""
    for row in range(2, sheet.max_row + 1):
        for col in columns_monetarias:
            cell = sheet.cell(row=row, column=col)
            cell.number_format = 'R$ #,##0.00'
        for col in columns_percentuais:
            cell = sheet.cell(row=row, column=col)
            cell.number_format = '0.00%'

def exportar_xlsx(data):
    """Exporta os resultados para um arquivo XLSX com formatação profissional e visual amigável"""
    
    vendas = data['vendas']
    matriz = data['matriz'] if data['matriz'] is not None else pd.DataFrame()
    full = data['full'] if data['full'] is not None else pd.DataFrame()
    max_date = data['max_date']
    
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    
    # 1. ABA RESUMO EXECUTIVO
    metricas_total = calcular_metricas(vendas, matriz, full, max_date, 180)
    
    resumo_data = [
        ['MÉTRICA DE DESEMPENHO', 'VALOR ATUAL (ÚLTIMOS 180 DIAS)'],
        ['Total de Pedidos Processados', metricas_total['vendas']],
        ['Faturamento Bruto de Produtos', metricas_total['faturamento_produtos']],
        ['Faturamento Total (c/ Fretes)', metricas_total['faturamento_total']],
        ['---', '---'],
        ['Total de Devoluções', metricas_total['devolucoes_vendas']],
        ['Taxa de Devolução Global', metricas_total['taxa_devolucao']],
        ['Faturamento Comprometido (Devoluções)', metricas_total['faturamento_devolucoes']],
        ['---', '---'],
        ['Perda Total Estimada', abs(metricas_total['perda_total'])],
        ['Perda Parcial (Custos Operacionais)', abs(metricas_total['perda_parcial'])],
        ['---', '---'],
        ['Análise de Saúde das Devoluções', 'Quantidade'],
        ['Devoluções Saudáveis (Reembolsadas)', metricas_total['saudaveis']],
        ['Devoluções Críticas (Prejuízo Total)', metricas_total['criticas']],
        ['Devoluções Neutras', metricas_total['neutras']],
    ]
    
    df_resumo = pd.DataFrame(resumo_data[1:], columns=resumo_data[0])
    df_resumo.to_excel(writer, sheet_name='Resumo Executivo', index=False)
    
    ws_resumo = writer.sheets['Resumo Executivo']
    aplicar_estilo_cabecalho(ws_resumo, 2)
    formatar_valores_excel(ws_resumo, [], [2]) # Taxa está na linha 7, mas formatar_valores_excel é por coluna
    # Ajuste manual de formatos para a aba resumo que é vertical
    for row in [3, 4, 8, 10, 11]:
        ws_resumo.cell(row=row, column=2).number_format = 'R$ #,##0.00'
    ws_resumo.cell(row=7, column=2).number_format = '0.00%'
    ajustar_largura_colunas(ws_resumo)

    # 2. ABA ANÁLISE TEMPORAL (JANELAS)
    janelas_list = []
    for janela in [30, 60, 90, 180]:
        m = calcular_metricas(vendas, matriz, full, max_date, janela)
        janelas_list.append({
            'Período': f'Últimos {janela} dias',
            'Vendas': m['vendas'],
            'Faturamento Prod.': m['faturamento_produtos'],
            'Devoluções': m['devolucoes_vendas'],
            'Taxa %': m['taxa_devolucao'],
            'Perda Total': abs(m['perda_total']),
            'Perda/Venda Média': abs(m['perda_total'] / m['vendas']) if m['vendas'] > 0 else 0
        })
    
    df_janelas = pd.DataFrame(janelas_list)
    df_janelas.to_excel(writer, sheet_name='Análise Temporal', index=False)
    ws_janelas = writer.sheets['Análise Temporal']
    aplicar_estilo_cabecalho(ws_janelas, 7)
    formatar_valores_excel(ws_janelas, [3, 6, 7], [5])
    ajustar_largura_colunas(ws_janelas)

    # 3. ABA QUALIDADE DOS DADOS
    qualidade = calcular_qualidade_arquivo(data)
    qualidade_rows = [
        ['INDICADOR DE QUALIDADE', 'PERCENTUAL DE FALHA'],
        ['Vendas: SKUs não identificados', qualidade['vendas']['sem_sku_pct'] / 100],
        ['Vendas: Datas ausentes', qualidade['vendas']['sem_data_pct'] / 100],
        ['Vendas: N.º de venda ausente', qualidade['vendas']['sem_numero_venda_pct'] / 100],
        ['Devoluções Matriz: Sem motivo informado', qualidade['matriz']['sem_motivo_pct'] / 100],
        ['Devoluções Matriz: Sem estado do produto', qualidade['matriz']['sem_estado_pct'] / 100],
        ['Devoluções Full: Sem motivo informado', qualidade['full']['sem_motivo_pct'] / 100],
        ['Devoluções Full: Sem estado do produto', qualidade['full']['sem_estado_pct'] / 100],
    ]
    pd.DataFrame(qualidade_rows[1:], columns=qualidade_rows[0]).to_excel(writer, sheet_name='Qualidade dos Dados', index=False)
    ws_qual = writer.sheets['Qualidade dos Dados']
    aplicar_estilo_cabecalho(ws_qual, 2)
    formatar_valores_excel(ws_qual, [], [2])
    ajustar_largura_colunas(ws_qual)

    # 4. ABA DADOS DE VENDAS (AMOSTRA)
    vendas_cols = ['Data da venda', 'N.º de venda', 'Título do anúncio', 'SKU', 'Quantidade', 'Preço unitário de venda de produtos', 'Tarifa de venda e impostos', 'Custo de envio']
    vendas_export = vendas[vendas_cols].head(2000) if all(c in vendas.columns for c in vendas_cols) else vendas.head(2000)
    vendas_export.to_excel(writer, sheet_name='Base de Vendas', index=False)
    ws_vendas = writer.sheets['Base de Vendas']
    aplicar_estilo_cabecalho(ws_vendas, len(vendas_export.columns))
    # Tentar identificar colunas monetárias na base de vendas
    monetarias_vendas = []
    for i, col in enumerate(vendas_export.columns):
        if any(term in col.lower() for term in ['preço', 'tarifa', 'custo', 'valor', 'faturamento']):
            monetarias_vendas.append(i + 1)
    formatar_valores_excel(ws_vendas, monetarias_vendas, [])
    ajustar_largura_colunas(ws_vendas)

    # 5. ABA DEVOLUÇÕES (CONSOLIDADO)
    devolucoes_raw = pd.concat([matriz, full], ignore_index=True)
    if not devolucoes_raw.empty:
        dev_cols = ['Data da venda', 'N.º de venda', 'Título do anúncio', 'SKU', 'Motivo da devolução', 'Estado do produto', 'Valor reembolsado', 'Custos de devolução']
        dev_export = devolucoes_raw[dev_cols].head(2000) if all(c in devolucoes_raw.columns for c in dev_cols) else devolucoes_raw.head(2000)
        dev_export.to_excel(writer, sheet_name='Base de Devoluções', index=False)
        ws_dev = writer.sheets['Base de Devoluções']
        aplicar_estilo_cabecalho(ws_dev, len(dev_export.columns))
        monetarias_dev = []
        for i, col in enumerate(dev_export.columns):
            if any(term in col.lower() for term in ['reembolsado', 'custo', 'valor', 'preço']):
                monetarias_dev.append(i + 1)
        formatar_valores_excel(ws_dev, monetarias_dev, [])
        ajustar_largura_colunas(ws_dev)

    writer.close()
    output.seek(0)
    return output
